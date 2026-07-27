"""
Data loader & cleaner for Laporan Keuangan Masjid Al Ikhlas.

Reads the raw ledger from Sheet1 of the source .xlsx (transaction log
from Jan 2019 - Jun 2026) and returns a clean pandas DataFrame plus
handy monthly / category aggregations used by the dashboard.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
# Raw ledger lives in Sheet1: header row 8, data rows 9..684 (1-indexed as
# seen in Excel), columns: NO, TANGGAL, KETERANGAN, MASUK, KELUAR, SALDO
# ---------------------------------------------------------------------------

SHEET_NAME = "Sheet1"
HEADER_ROW = 7          # 0-indexed -> Excel row 8

MONTHS_ID = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
    "juli": 7, "agustus": 8, "september": 9, "oktober": 10,
    "november": 11, "desember": 12,
}


def _parse_indo_date(value):
    """Parse either a datetime already-parsed by openpyxl, or an Indonesian
    text date like '20 Januari 2019' / '25 Februari 2019 ' into a Timestamp.
    Returns pd.NaT if it can't be parsed."""
    if value is None:
        return pd.NaT
    if isinstance(value, (datetime, pd.Timestamp)):
        return pd.Timestamp(value)
    if isinstance(value, str):
        s = value.strip().lower()
        s = re.sub(r"\s+", " ", s)
        m = re.match(r"(\d{1,2})\s+([a-z]+)\s+(\d{4})", s)
        if m:
            day, month_name, year = m.groups()
            month = MONTHS_ID.get(month_name)
            if month:
                try:
                    return pd.Timestamp(year=int(year), month=month, day=int(day))
                except ValueError:
                    return pd.NaT
    return pd.NaT


def _to_number(value):
    if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = re.sub(r"[^0-9,.-]", "", value)
        cleaned = cleaned.replace(".", "").replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


# Keyword -> category, checked in order (first match wins).
CATEGORY_RULES = [
    (r"gaji|thr\b", "Gaji Marbot / Staf"),
    (r"listrik|token li|listik|kabel|service ac|pengerjaan ac|perbaikan|penyambungan", "Listrik & Perawatan"),
    (r"\btpa\b|sanlat|satlat", "TPA"),
    (r"majelis ta'?lim|majelis talim", "Majelis Ta'lim"),
    (r"kotak amal|kencleng", "Kotak Amal / Kencleng"),
    (r"rekapan infaq|infaq.*rekening|infaq.*qris|qris|dari jamaah via transfer", "Infaq Rekening/QRIS"),
    (r"alokasi dana.*rw|dana dari rw|bantuan dari rw|bantuan dari rt", "Dana RW/RT"),
    (r"khotib|ustadz|imam|ceramah|tarawih|tarhib", "Infaq Ustadz/Khotib/Imam"),
    (r"konsumsi|snack|rapat dkm", "Konsumsi & Kegiatan"),
    (r"karpet|renovasi|\batap\b|speaker|sound|dispenser|pembangunan", "Renovasi & Aset Masjid"),
    (r"hadiah|lomba|maulid|pawai|obor|taawun|pengobatan", "Kegiatan & Sosial"),
    (r"sumbangan|bantuan|infaq ", "Infaq & Sumbangan Lain"),
    (r"saldo awal", "Saldo Awal"),
]


def _categorize(keterangan: str) -> str:
    if not keterangan:
        return "Lainnya"
    text = keterangan.lower()
    for pattern, label in CATEGORY_RULES:
        if re.search(pattern, text):
            return label
    return "Lainnya"


CANONICAL_SHEET_NAME = "Ledger"


def load_ledger(xlsx_path: str | Path) -> pd.DataFrame:
    """Load & clean the ledger, auto-detecting the file format.

    Supports two formats:
    - The original monthly-report layout (Sheet1, header at Excel row 8) —
      used for the raw files the DKM hands you each month.
    - The app's own "canonical" export layout (sheet named "Ledger", header
      on row 1, includes a KATEGORI column) — used once transactions have
      been persisted permanently via the dashboard's "Simpan Permanen"
      button, so re-loading the saved file round-trips correctly.
    """
    try:
        sheet_names = pd.ExcelFile(xlsx_path, engine="openpyxl").sheet_names
    except Exception:
        sheet_names = []

    if CANONICAL_SHEET_NAME in sheet_names:
        return _load_canonical_format(xlsx_path)
    return _load_raw_monthly_report(xlsx_path)


def _load_raw_monthly_report(xlsx_path: str | Path) -> pd.DataFrame:
    """Parse the original monthly-report ledger: Sheet1, header row 8,
    columns NO, TANGGAL, KETERANGAN, MASUK, KELUAR, SALDO."""
    raw = pd.read_excel(
        xlsx_path,
        sheet_name=SHEET_NAME,
        header=HEADER_ROW,
        engine="openpyxl",
    )
    raw = raw.iloc[:, :6]
    raw.columns = ["no", "tanggal", "keterangan", "masuk", "keluar", "saldo"]

    # Keep only real transaction rows: must have a keterangan and (masuk or keluar or saldo)
    raw = raw[raw["keterangan"].notna()].copy()
    raw = raw[~raw["keterangan"].astype(str).str.contains("total saldo", case=False, na=False)]

    raw["tanggal"] = raw["tanggal"].apply(_parse_indo_date)
    raw["masuk"] = raw["masuk"].apply(_to_number)
    raw["keluar"] = raw["keluar"].apply(_to_number)
    raw["saldo"] = raw["saldo"].apply(_to_number)
    raw["keterangan"] = raw["keterangan"].astype(str).str.strip()

    raw = raw[raw["tanggal"].notna()].copy()
    raw = raw.sort_values("tanggal").reset_index(drop=True)

    raw["kategori"] = raw["keterangan"].apply(_categorize)
    raw["net"] = raw["masuk"] - raw["keluar"]
    raw["bulan"] = raw["tanggal"].dt.to_period("M").dt.to_timestamp()
    raw["tahun"] = raw["tanggal"].dt.year

    return raw[["no", "tanggal", "bulan", "tahun", "keterangan", "kategori",
                "masuk", "keluar", "net", "saldo"]]


def _load_canonical_format(xlsx_path: str | Path) -> pd.DataFrame:
    """Parse the app's own export layout: sheet "Ledger", header row 1,
    columns NO, TANGGAL, KETERANGAN, KATEGORI, MASUK, KELUAR, SALDO."""
    raw = pd.read_excel(
        xlsx_path,
        sheet_name=CANONICAL_SHEET_NAME,
        header=0,
        engine="openpyxl",
    )
    raw.columns = [str(c).strip().lower() for c in raw.columns]

    raw = raw[raw["keterangan"].notna()].copy()
    raw = raw[~raw["keterangan"].astype(str).str.strip().str.upper().eq("TOTAL")]

    raw["tanggal"] = raw["tanggal"].apply(_parse_indo_date)
    raw["masuk"] = raw["masuk"].apply(_to_number)
    raw["keluar"] = raw["keluar"].apply(_to_number)
    raw["saldo"] = raw["saldo"].apply(_to_number)
    raw["keterangan"] = raw["keterangan"].astype(str).str.strip()

    if "kategori" in raw.columns:
        raw["kategori"] = raw["kategori"].fillna("").astype(str).str.strip()
    else:
        raw["kategori"] = ""
    blank_kategori = raw["kategori"] == ""
    raw.loc[blank_kategori, "kategori"] = raw.loc[blank_kategori, "keterangan"].apply(_categorize)

    raw = raw[raw["tanggal"].notna()].copy()
    raw = raw.sort_values(["tanggal"], kind="stable").reset_index(drop=True)

    raw["net"] = raw["masuk"] - raw["keluar"]
    raw["bulan"] = raw["tanggal"].dt.to_period("M").dt.to_timestamp()
    raw["tahun"] = raw["tanggal"].dt.year
    raw["no"] = range(1, len(raw) + 1)

    return raw[["no", "tanggal", "bulan", "tahun", "keterangan", "kategori",
                "masuk", "keluar", "net", "saldo"]]


def monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    g = (
        df.groupby("bulan")
        .agg(masuk=("masuk", "sum"), keluar=("keluar", "sum"))
        .reset_index()
        .sort_values("bulan")
    )
    g["net"] = g["masuk"] - g["keluar"]
    g["saldo_akhir_bulan"] = df.groupby("bulan")["saldo"].last().values
    return g


def category_summary(df: pd.DataFrame, kind: str = "keluar") -> pd.DataFrame:
    col = "keluar" if kind == "keluar" else "masuk"
    g = (
        df[df[col] > 0]
        .groupby("kategori")[col]
        .sum()
        .reset_index()
        .sort_values(col, ascending=False)
    )
    g.columns = ["kategori", "total"]
    return g


def append_transaction(
    df: pd.DataFrame,
    tanggal,
    keterangan: str,
    kategori: str,
    masuk: float = 0.0,
    keluar: float = 0.0,
) -> pd.DataFrame:
    """Return a new ledger DataFrame with one extra transaction appended
    at the end, saldo carried forward from the last recorded balance.

    Note: the original ledger's SALDO column contains a handful of manual
    adjustments that don't perfectly match cumulative masuk-keluar (typical
    of a hand-kept spreadsheet), so existing rows are left untouched and
    only the new row's saldo is derived from the last known balance. New
    transactions are always appended after the last row, regardless of the
    date entered, so historical saldo figures are never recalculated.
    """
    tanggal = pd.Timestamp(tanggal)
    last_saldo = float(df["saldo"].iloc[-1]) if not df.empty else 0.0
    net = float(masuk) - float(keluar)

    new_row = {
        "no": (int(df["no"].max()) + 1) if not df.empty else 1,
        "tanggal": tanggal,
        "bulan": tanggal.to_period("M").to_timestamp(),
        "tahun": tanggal.year,
        "keterangan": keterangan.strip(),
        "kategori": kategori,
        "masuk": float(masuk),
        "keluar": float(keluar),
        "net": net,
        "saldo": last_saldo + net,
    }

    out = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    return out


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Ledger") -> bytes:
    """Export the ledger to a nicely formatted .xlsx file (bytes buffer),
    ready for st.download_button or writing to disk."""
    import io

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    export_cols = ["no", "tanggal", "keterangan", "kategori", "masuk", "keluar", "saldo"]
    export_df = df[export_cols].rename(
        columns={
            "no": "NO",
            "tanggal": "TANGGAL",
            "keterangan": "KETERANGAN",
            "kategori": "KATEGORI",
            "masuk": "MASUK",
            "keluar": "KELUAR",
            "saldo": "SALDO",
        }
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="1F8A70", end_color="1F8A70", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(export_cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        widths = {"NO": 6, "TANGGAL": 14, "KETERANGAN": 45, "KATEGORI": 24,
                  "MASUK": 15, "KELUAR": 15, "SALDO": 16}
        for i, col in enumerate(export_cols, start=1):
            header = export_df.columns[i - 1]
            ws.column_dimensions[get_column_letter(i)].width = widths.get(header, 16)

        for row in range(2, len(export_df) + 2):
            ws.cell(row=row, column=2).number_format = "dd mmm yyyy"
            for col in (5, 6, 7):
                ws.cell(row=row, column=col).number_format = "#,##0"

        total_row = len(export_df) + 2
        ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=float(export_df["MASUK"].sum())).number_format = "#,##0"
        ws.cell(row=total_row, column=6, value=float(export_df["KELUAR"].sum())).number_format = "#,##0"
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.freeze_panes = "A2"

    return buffer.getvalue()


if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else "data/laporan_keuangan.xlsx"
    df = load_ledger(path)
    print(f"Loaded {len(df)} transactions from {df['tanggal'].min().date()} to {df['tanggal'].max().date()}")
    print(f"Total masuk : {df['masuk'].sum():,.0f}")
    print(f"Total keluar: {df['keluar'].sum():,.0f}")
    print(f"Saldo akhir : {df['saldo'].iloc[-1]:,.0f}")
    print("\nKategori (keluar):")
    print(category_summary(df, "keluar").to_string(index=False))
    print("\nKategori (masuk):")
    print(category_summary(df, "masuk").to_string(index=False))
    print("\nRingkasan bulanan (5 terakhir):")
    print(monthly_summary(df).tail(5).to_string(index=False))
